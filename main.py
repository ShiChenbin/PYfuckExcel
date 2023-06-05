# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.
import sys
try:
    import pymysql
    from openpyxl import load_workbook
    import tushare as ts
    from datetime import datetime, timedelta
    import os
    import re
    import pandas as pd
    import sqlite3


    current_dir = os.path.dirname(os.path.realpath(sys.argv[0]))  # 获取当前脚本所在的目录路径 os.path.abspath(__file__)
    # print(current_dir)
    # 建立汇富联合（hf）数据库连接
    db_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'hf.db')
    conn = sqlite3.connect(db_file)
    print(db_file)
    # 创建操作数据库的游标
    cursor = conn.cursor()

    # 调用ts的api
    ts.set_token('c81f5a49a33c8e20f7b436f2e6ab83942d5343b16e8debb7c235c55a')
    pro = ts.pro_api()

    # 读取excel
    relative_path = 'zhengquan2023.xlsx'  # 相对于当前目录的路径
    file_path = os.path.join(current_dir, relative_path)  # 构建完整的文件路径
    print(file_path)
    zq_excel_wb = load_workbook(file_path, data_only=True)

    # 选择表格1
    # zq_ws1 = zq_excel_wb['日盈5中金']

    # 所有表名字
    # all_sheet_name = zq_excel_wb.sheetnames
    # print(all_sheet_name)

    # 循环遍历所有表格
    for sheet_name in zq_excel_wb.sheetnames:
        zq_ws = zq_excel_wb[sheet_name]
        print(sheet_name)
        # 循环遍历每一行，从第3行开始，跳过表头
        for index in range(3, zq_ws.max_row + 1):
            # 获取C列（股票代码）
            stock_code = zq_ws.cell(row=index, column=3).value
            print(stock_code)
            # 如果C列没有数据，跳出循环
            if not stock_code:
                break

            judge_stock_code = str(stock_code)
            if re.match(r"^\d{6}$", judge_stock_code):
                print("stock_code是股票代码")
            else:
                print("stock_code不是股票代码")
                break

            #if not isinstance(stock_code, str) or not stock_code.isdigit() or len(stock_code) != 6:
            #    continue  # 跳过当前循环

            # 股票代码应区别于证券查找的代码

            # 获取股票的出借时间格式化为目标时间字符串,这里的stock_lend_time是时间格式的
            stock_lend_time = zq_ws.cell(row=index, column=2).value
            # print(stock_lend_time)  # 输出：2022-12-28 00:00:00
            if not isinstance(stock_lend_time, datetime):
                continue  # 跳过循环

            # 将日期时间对象转换为字符串格式
            find_lend_time = stock_lend_time.strftime("%Y%m%d")
            print(find_lend_time)  # 输出：20221228

            # 判断加减的天数是不是数字，如果是数字则可以进行处理，不是则跳出当前行处理
            if not isinstance(zq_ws.cell(row=index, column=9).value, (int, float)) or not str(zq_ws.cell(row=index, column=9).value).isnumeric():
                break  # 跳出循环
            # 根据表格所提供的借入时间和借入天数
            days_to_add = int(zq_ws.cell(row=index, column=9).value) - 1

            # print(days_to_add)

            stock_end_time = stock_lend_time + timedelta(days=days_to_add)
            stock_end_time = stock_end_time.strftime('%Y/%m/%d') # 格式处理
            zq_ws.cell(row=index, column=12).value = stock_end_time

            # 选择最近的交易日作为还券日期
            if zq_ws.cell(row=index, column=11).value is None:
                # 生成日期范围，只包含工作日
                date_range = pd.bdate_range(start=stock_lend_time, end=stock_end_time)

                # 选择最近的工作日
                stock_huanquan_time = date_range[-1]
                stock_huanquan_time = stock_huanquan_time.strftime('%Y/%m/%d')
                zq_ws.cell(row=index, column=11).value = stock_huanquan_time

            # 根据stock_code = symbol的值，去查询
            # 查询数据库获取ts_code值
            ts_code_query = f"SELECT ts_code FROM stock_info WHERE symbol = '{stock_code}'"
            cursor.execute(ts_code_query)
            result_ts_code = cursor.fetchone()

            # 获取ts_code值
            get_ts_code = result_ts_code[0] if result_ts_code else ''


            # 查询股票当日收盘价
            df = pro.daily(ts_code=get_ts_code, start_date=find_lend_time, end_date=find_lend_time)
            print(df)

            # 查询股票名称
            stock_name_query = f"SELECT name FROM stock_info WHERE symbol = '{stock_code}'"
            cursor.execute(stock_name_query)
            result_name = cursor.fetchone()

            # 获取股票名称
            stock_name = result_name[0] if result_name else ''

            # 将股票名称填入表中的第四列
            zq_ws.cell(row=index, column=4).value = stock_name


            if not df.empty:
                close_price = df.iloc[0]['close']

                # 更新excel当中的收盘价
                zq_ws.cell(row=index, column=6).value = close_price

                # 根据输入的股数计算市值
                zq_ws.cell(row=index, column=7).value = close_price * zq_ws.cell(row=index, column=5).value

                # 市值
                all_value = zq_ws.cell(row=index, column=7).value

                # 计算预计券息
                qx_result = all_value * zq_ws.cell(row=index, column=13).value * zq_ws.cell(row=index, column=8).value / 360
                qx_result_rounded = round(qx_result, 2)
                zq_ws.cell(row=index, column=14).value = qx_result_rounded

                # 更新数据库中的close项
                # update_query = f"UPDATE stock_info SET close = {close_price} WHERE symbol = '{stock_code}'"
                # cursor.execute(update_query)
                # conn.commit()




    # 关闭游标和数据库连接
    cursor.close()
    conn.close()

    # 保存修改后的Excel文件
    zq_excel_wb.save(file_path)
    zq_excel_wb.close()

except Exception as e:
    print(f"Error: {e}")

input("Press Enter to exit...")
sys.exit()

# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    print("successful end")
